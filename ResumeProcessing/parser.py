#!/usr/bin/env python
# coding: utf-8


# importing all the libraries
import io
import os
import re
import pandas as pd
import nltk
from nltk.tokenize import sent_tokenize, word_tokenize, TreebankWordTokenizer, WordPunctTokenizer, MWETokenizer
# import PyPDF2
# for pdf to txt
import pdfminer
from pdfminer.pdfinterp import PDFResourceManager, PDFPageInterpreter
from pdfminer.layout import LAParams
from pdfminer.converter import TextConverter
from io import StringIO
from pdfminer.pdfpage import PDFPage
import openpyxl as px
import docx2txt
import constants as cs
import string
# import utils
import pprint
from spacy.matcher import matcher
import multiprocessing as mp
import warnings
from nltk.corpus import stopwords
from nltk.stem import PorterStemmer
from nltk.tokenize import word_tokenize
from transformers import XLNetTokenizer
import torch
import tensorflow as tf
from transformers import AutoTokenizer, AutoModelForTokenClassification
from transformers import pipeline

# from constants import STOPWORDS
warnings.filterwarnings('ignore')


titles = ["Name", "Email", "Phone", "Skills","Matched"]
resume_data={}


resume_file=input('Enter resume: ')
print(resume_file)


# Extract txt from pdf
# pdf_path="code\Resume.pdf"
def pdf_to_txt(pdf_path):
    resource_manager = PDFResourceManager(caching=True)
    
    # create a string object that will contain the final text the representation of the pdf. 
    out_text = StringIO()
    codec = 'utf-8'
    laParams = LAParams()
    
    # Create a TextConverter Object:
    text_converter = TextConverter(resource_manager, out_text, laparams=laParams)
    fp = open(pdf_path, 'rb')
    #Create a PDF interpreter object

    interpreter = PDFPageInterpreter(resource_manager, text_converter)
    
    # We are going to process the content of each page of the original PDF File    
    for page in PDFPage.get_pages(fp, pagenos=set(), maxpages=0, password="", caching=True, check_extractable=True):
        interpreter.process_page(page)

    
    # Retrieve the entire contents of the “file” at any time 
    text = out_text.getvalue()

    # Closing all the ressources we previously opened

    fp.close()
    text_converter.close()
    out_text.close()
    
    return text
# retext = pdf_to_txt(pdf_path)
# print(pdf_to_txt(pdf_path))

# Extract txt from
# docx_path="Resume.docx"
def docx_to_txt(docx_path):
    temp = docx2txt.process(docx_path)
    text = [line.replace('	', ' ') for line in temp.split('') if line]
    return ' '.join(text)
# print(docx_to_txt(docx_path))

split_tup = os.path.splitext(resume_file)
print("hello",split_tup)
  
# extract the file name and extension
file_name = split_tup[0]
file_extension = split_tup[1]
  
print("File Name: ", file_name)
print("File Extension: ", file_extension)


# detect file extension and call above functions accordingly
def extract_text(file_path, extension):
    '''
    Wrapper function to detect the file extension and call text extraction function accordingly

    :param file_path: path of file of which text is to be extracted
    :param extension: extension of file `file_name`
    '''
    text = ''
    if extension == '.pdf':
        for page in pdf_to_txt(file_path):
            text += '' + page
    elif extension == '.docx' or extension == '.doc':
        text = docx_to_txt(file_path)
    return text
Text= extract_text(resume_file,file_extension)
print(Text)


# Tokenization Of the Extracted Text

nltk.download('all')
# Word Tokenize

print("word_tokenize",word_tokenize(Text))

tokenizer = TreebankWordTokenizer()
print("TreebankWordTokenizer",tokenizer.tokenize(Text))

tokenizer = WordPunctTokenizer()
print("WordPunctTokenizer",tokenizer.tokenize(Text) )

from nltk.tokenize import MWETokenizer
mtokenizer = MWETokenizer([('athale20comp@studen', 't.mes.ac.in'), ('+91', '-', '8605131403')],separator='')
aftertoken = mtokenizer.tokenize(Text.split())
print("MWETokenizer",aftertoken)
aftertoken=(' '.join(aftertoken))
print(aftertoken)



from nltk.tokenize import sent_tokenize
print(sent_tokenize(Text))


# Filteration


def  filter_text(Text,lowerFlag=False,upperFlag=False,numberFlag=False,htmlFlag=False,urlFlag=False,punctFlag=False,spaceFlag=False,charFlag=False,bulletFlag=False,hashtagFlag=False,emojiFlag=False):
    if lowerFlag:
      Text = Text.lower()

    if upperFlag:
      Text = Text.upper()

    if numberFlag:
      import re
      Text = re.sub(r"\d+", '', Text)

    if htmlFlag:
      import re
      Text = re.sub(r'<[^>]*>', '', Text)

    if urlFlag:
      import re
      Text = re.sub(r'(https?|ftp|www)\S+', '', Text)

    if punctFlag:
      import re
      import string
      exclist = string.punctuation #removes [!”#$%&’()*+,-./:;<=>?@[\]^_`{|}~]
      # remove punctuations and digits from oldtext
      table_ = Text.maketrans('', '', exclist)
      Text = Text.translate(table_)

    if spaceFlag:
      import re
      Text = re.sub(' +'," ",Text).strip()

    if hashtagFlag:
      import re
      Text = re.sub(' +',"#",Text).strip()
      
    if charFlag:
      import re
      Text = re.sub('', ' ', Text)
      
    if bulletFlag:
      import re
      Text = re.sub('●','',Text)
      Text = re.sub('★','',Text)
      Text = re.sub('|','',Text)
      # Text = re.sub(([0-9|A-Z]+[\.|\)]|)\s+,'', Text)
    
    if emojiFlag:
        inText = re.sub("["
                           u"\U0001F600-\U0001F64F"  # Emojis in Emoticons
                           u"\U0001F300-\U0001F5FF"  # Emojis in Miscellaneous Symbols and Pictographs
                           u"\U0001F680-\U0001F6FF"  # Emojis in Transport and Map Symbols
                           u"\U0001F700-\U0001F77F"  # Emojis in Alphabetic Presentation Forms
                           u"\U0001F780-\U0001F7FF"  # Emojis in Arabic Presentation Forms-A
                           u"\U0001F800-\U0001F8FF"  # Emojis in Supplemental Arrows-C
                           u"\U0001F900-\U0001F9FF"  # Emojis in Supplemental Symbols and Pictographs
                           u"\U0001FA00-\U0001FA6F"  # Emojis in Chess Symbols
                           u"\U0001FA70-\U0001FAFF"  # Emojis in Symbols and Pictographs Extended-A
                           u"\U0001F004-\U0001F0CF"  # Additional Emojis
                           u"\U0001F0D0-\U0001F0FF"  # Playing Cards
                           u"\U0001F10D-\U0001F10F"  # Emoticons (Hearts)
                           u"\U0001F170-\U0001F251"  # Emoticons (Enclosed Characters)
                           "]+", '', inText, flags=re.UNICODE)

    return Text



filteredtxt = filter_text(aftertoken)
print(filteredtxt)


# Stopwords


# from nltk.corpus import stopwords
stop_words = set(stopwords.words("english"))
print(stop_words)




mtokenizer = MWETokenizer([('athale20comp@studen', 't.mes.ac.in'), ('+91', '-', '8605131403')],separator='')
word_tokens = mtokenizer.tokenize(filteredtxt.split())

# print(ex_text)

stop = [w for w in stop_words if w in word_tokens]
print("StopWords recognized in the given sentence:", stop,"")

filtered_sentence = [w for w in word_tokens if not w in stop_words]
filtered_sentence = []

for w in word_tokens:
    if w not in stop_words:
        filtered_sentence.append(w)

print("After removing the recognized stopwords, the Tokens of sentence is:", filtered_sentence)
filtered_data=(' '.join(filtered_sentence))
print(filtered_data)


# Stemming or Lemmatization


import spacy
nlp = spacy.load('en_core_web_sm')

# Create a Doc object
doc = nlp(filtered_data)

# Create list of tokens from given string
tokens = []
for token in doc:
	tokens.append(token)

print(tokens)
#> [the, bats, saw, the, cats, with, best, stripes, hanging, upside, down, by, their, feet]

lemmatized_sentence = " ".join([token.lemma_ for token in doc])

print(lemmatized_sentence)
#> the bat see the cat with good stripe hang upside down by -PRON- foot



def extract_email(text):
    '''
    Helper function to extract email id from text

    :param text: plain text extracted from resume file
    '''
    email = re.findall("([^@|\s]+@[^@]+\.[^@|\s]+)", text)
    if email:
        try:
            return email[0].split()[0].strip(';')
        except IndexError:
            return None
print("E-mail: ",extract_email(filtered_data))
resume_data[titles[1]]=extract_email(filtered_data)




def extract_mobile_number(text):
    '''
    Helper function to extract mobile number from text

    :param text: plain text extracted from resume file
    :return: string of extracted mobile numbers
    '''
    # Found this complicated regex on : https://zapier.com/blog/extract-links-email-phone-regex/
    phone = re.findall(re.compile(r'(?:(?:\+?([1-9]|[0-9][0-9]|[0-9][0-9][0-9])\s*(?:[.-]\s*)?)?(?:\(\s*([2-9]1[02-9]|[2-9][02-8]1|[2-9][02-8][02-9])\s*\)|([0-9][1-9]|[0-9]1[02-9]|[2-9][02-8]1|[2-9][02-8][02-9]))\s*(?:[.-]\s*)?)?([2-9]1[02-9]|[2-9][02-9]1|[2-9][02-9]{2})\s*(?:[.-]\s*)?([0-9]{4})(?:\s*(?:#|x\.?|ext\.?|extension)\s*(\d+))?'), text)
    if phone:
        number = ''.join(phone[0])
        if len(number) > 10:
            return '+' + number
        else:
            return number
print("Mobile No.: ",extract_mobile_number(filtered_data))
resume_data[titles[2]]=extract_mobile_number(filtered_data)




def list_to_string(lst):
    # Convert the list to a string representation
    return ', '.join(map(str, lst))
noun_chunks=[]
extracted_skills=[]
file='skills.csv'
new_nlp = nlp(filtered_data)
def extract_skills(nlp_text, noun_chunks):
    '''
    Helper function to extract skills from spacy nlp text

    :param nlp_text: object of `spacy.tokens.doc.Doc`
    :param noun_chunks: noun chunks extracted from nlp text
    :return: list of skills extracted
    '''
    tokens = [token.text for token in nlp_text if not token.is_stop]
    data = pd.read_csv(os.path.join(os.path.dirname(file), r'ResumeProcessing\skills.csv'))
    skills = list(data.columns.values)
    skillset = []
    # check for one-grams
    for token in tokens:
        if token.lower() in skills:
            skillset.append(token)
    
    # check for bi-grams and tri-grams
    for token in noun_chunks:
        token = token.text.lower().strip()
        if token in skills:
            skillset.append(token)
    extracted_skills=[i.capitalize() for i in set([i.lower() for i in skillset])]

    return [i.capitalize() for i in set([i.lower() for i in skillset])]
extract_skills(new_nlp,noun_chunks)
extracted_skills=extract_skills(new_nlp,noun_chunks)
print(extracted_skills)
resume_data[titles[3]]=list_to_string(extracted_skills)




# Matching the extracted skills with our requirements


file='ResumeProcessing\required.csv'
new_nlp = nlp(filtered_data)
from difflib import SequenceMatcher

def match_skills(skills):
    '''
    Helper function to extract skills from spacy nlp text

    :param nlp_text: object of `spacy.tokens.doc.Doc`
    :param noun_chunks: noun chunks extracted from nlp text
    :return: list of skills extracted
    '''
    # tokens = [token.text for token in nlp_text if not token.is_stop]
    # skills=[]
    data = pd.read_csv(os.path.join(os.path.dirname(file),  r'ResumeProcessing\required.csv')) 
    required = list(data.columns.values)
    print("Skills in Resume: ",skills)
    print("Required skills are: ",required)
    matched = []
    # check for one-grams
    for skill in skills:
        if skill.lower() in required:
            matched.append(skill)
    print("Matched skills are: ",matched)
    res = len(set(required) and set(matched)) / float(len(set(required) or set(matched))) * 100
    print("The percentage of skills matched is: ",res)

    from math import sqrt

    def dot_product(vec1, vec2):
        return sum(x * y for x, y in zip(vec1, vec2))

    def vector_length(vec):
        return sqrt(sum(x * x for x in vec))

    def cosine_similarity(vec1, vec2):
        if vector_length(vec1) == 0 or vector_length(vec2) == 0:
            return 0.0
        return dot_product(vec1, vec2) / (vector_length(vec1) * vector_length(vec2))

# Example usage:
    required_skills = ["Python", "Java", "SQL", "Machine Learning", "Communication"]
    candidate_skills = ["Python", "SQL", "Data Analysis"]

# Convert skills to binary vectors
    required_skills_vector = [1 if skill in required_skills else 0 for skill in required_skills]
    candidate_skills_vector = [1 if skill in candidate_skills else 0 for skill in required_skills]

    similarity_percentage = cosine_similarity(required_skills_vector, candidate_skills_vector) * 100
    print("Candidate's skill fit percentage (using cosine similarity):", similarity_percentage)

    return res






def extract_education(nlp_text):
    '''
    Helper function to extract education from spacy nlp text

    :param nlp_text: object of `spacy.tokens.doc.Doc`
    :return: tuple of education degree and year if year if found else only returns education degree
    '''
    EDUCATION         = [
                    'CBSE', 'ICSE', 'X', 'XII', 'BE', 'B.E.', 'B.E', 'BS', 'B.S', 'ME', 'M.E', 'M.E.', 'MS', 'M.S', 'BTECH', 'MTECH', 
                    'SSC', 'HSC','B.Tech','BTech'
                    ]
    
    STOPWORDS         = set(stopwords.words('english'))
    
    YEAR              = r'(((20|19)(\d{2})))'
    
    edu = {}
    # Extract education degree
    for index, text in enumerate(nlp_text):
        for tex in text.split():
            # tex = re.sub(r'[?|$|.|!|,]', r'', tex)
            if tex.upper() in EDUCATION:
                edu[tex] = text + nlp_text[index + 1]

    # Extract year
    education = []
    for key in edu.keys():
        year = re.search(re.compile(YEAR), edu[key])
        if year:
            education.append((key, ''.join(year.group(0))))
        else:
            education.append(key)
    return education
print("Education Qualification: ",extract_education(Text))


tokenizer2 = AutoTokenizer.from_pretrained("Davlan/distilbert-base-multilingual-cased-ner-hrl")
model2 = AutoModelForTokenClassification.from_pretrained("Davlan/distilbert-base-multilingual-cased-ner-hrl")
nlp2 = pipeline("ner", model=model2, tokenizer=tokenizer2, aggregation_strategy="max")
ner_results2 = nlp2(filtered_data)

print(ner_results2)

person_names = [entity['word'] for entity in ner_results2 if entity['entity_group'] == 'PER']


resume_data[titles[0]]=person_names[0]
resume_data[titles[4]]= match_skills(extracted_skills)
# import openpyxl
from openpyxl import Workbook, load_workbook
from datetime import datetime
import pandas as pd

import openpyxl as px

# Specify the Excel file path
excel_file = r"ResumeProcessing\Extracted Data\Extracted.xlsx"

# Load the existing Excel file
try:
    wb = px.load_workbook(excel_file)
except FileNotFoundError:
    # If the file doesn't exist, create a new workbook
    wb = px.Workbook()

# # Select the worksheet where you want to append data (or create a new one)
# ws = wb.active  # You can also select a specific sheet by name: workbook["SheetName"]

# Get the current timestamp
timestamp = datetime.now().strftime("%Y-%m-%d")
# setect the current active sheet
ws = wb.active
# check if the current sheet has same timestamp
if ws.title == timestamp:
    df = pd.DataFrame.from_dict([resume_data])
    
else:
    # Create a new sheet with the timestamp as the name
    ws = wb.create_sheet(title=timestamp, index=0)
    ws["A1"] = 'Email Id'
    ws["B1"] = 'Mobile No.'
    ws["C1"] = 'Skills'
    ws["D1"] = 'Name'
    ws["E1"] = 'Matched'
    df = pd.DataFrame.from_dict([resume_data])
  
      
for index, row in df.iterrows():
    ws.append(row.tolist())

wb.save(excel_file)




# ner using xlnet testing 



# import torch
# from transformers import XLNetForSequenceClassification, XLNetTokenizer, AdamW
# from torch.utils.data import DataLoader, Dataset
# import json

# # Load your JSON dataset
# with open("traindata3.json", "r", encoding="utf-8") as json_file:
#     data = json.load(json_file)





# def process_label_and_points(label, points):
#     # Initialize an empty list to store the extracted entity texts
#     entity_texts = []
    
#     # Loop through each point dictionary in the list
#     for point in points:
#         start = point['start']
#         end = point['end']
        
#         # Extract the labeled entity text from the content
#         entity_text = content[start:end]
        
#         # Append the entity text to the list
#         entity_texts.append(entity_text)
    
#     # Combine the extracted entity texts if needed (e.g., join with a space)
#     combined_entity_text = ' '.join(entity_texts)
    
#     return combined_entity_text

# # Initialize the XLNet tokenizer
# tokenizer = XLNetTokenizer.from_pretrained('xlnet-base-cased')

# # Tokenize and pad/truncate input sequences
# input_ids = []
# attention_masks = []

# for text in Text:
#     encoded_text = tokenizer.encode_plus(
#         text,
#         add_special_tokens=True,
#         max_length=128,  # Adjust the max length as needed
#         padding='max_length',
#         return_attention_mask=True,
#         return_tensors='pt'
#     )

#     input_ids.append(encoded_text['input_ids'])
#     attention_masks.append(encoded_text['attention_mask'])

# # Convert input_ids and attention_masks to tensors
# input_ids = torch.cat(input_ids, dim=0)
# attention_masks = torch.cat(attention_masks, dim=0)

# # Encode the target labels (NER tags) into tensors or other suitable format
# # You may need to use label encoders or other methods based on your model's requirements


# import torch
# from torch.utils.data import Dataset

# class NERDataset(Dataset):
#     def __init__(self, input_ids, attention_masks, target_labels):
#         """
#         Args:
#             input_ids (list of torch.Tensor): List of input sequences as tensors.
#             attention_masks (list of torch.Tensor): List of attention masks as tensors.
#             target_labels (list of list of str): List of target labels for NER, where each list contains labels for one example.
#         """
#         self.input_ids = input_ids
#         self.attention_masks = attention_masks
#         self.target_labels = target_labels

#     def __len__(self):
#         return len(self.input_ids)

#     def __getitem__(self, idx):
#         sample = {
#             'input_ids': self.input_ids[idx],
#             'attention_mask': self.attention_masks[idx],
#             'target_labels': self.target_labels[idx],
#         }
#         return sample



# from transformers import XLNetForTokenClassification, Trainer, TrainingArguments
# label_list = ['College Name', 'Designation', 'Name', 'Companies worked at', 'Email Address', 'Location', 'Skills']
# target_labels = [
#     ["O", "O", "B-ORG", "I-ORG", "O", ],  # Labels for the first example
#     ["B-PER", "I-PER", "O", "B-LOC", "O", ]  # Labels for the second example
#     # Add labels for more examples...
# ]

# # Calculate the number of unique labels
# num_labels = len(label_list)# Load the pre-trained XLNet model for token classification
# model = XLNetForTokenClassification.from_pretrained('xlnet-base-cased', num_labels=num_labels)  # Define num_labels accordingly

# # Define training arguments (batch size, learning rate, etc.)
# training_args = TrainingArguments(
#     output_dir='./ner_model',
#     per_device_train_batch_size=16,
#     num_train_epochs=3,
#     learning_rate=2e-5,
# )

# # Initialize the Trainer
# trainer = Trainer(
#     model=model,
#     args=training_args,
#     train_dataset=NERDataset(input_ids, attention_masks, target_labels),  # Define your custom dataset class
# )

# # Start training
# trainer.train()



# xlnet-base-cased
# MODEL_NAME = 'xlnet-base-cased'

# XLNetTokenizer.from_pretrained(MODEL_NAME, return_tensors="pt").save_pretrained("./{}_tokenizer".format(MODEL_NAME))

# # just in case if there is no TF/Keras file provided in the model
# # we can just use `from_pt` and convert PyTorch to TensorFlow
# try:
#     print('try downloading TF weights')
#     model = TFXLNetModel.from_pretrained(MODEL_NAME)
# except:
#     print('try downloading PyTorch weights')
#     model = TFXLNetModel.from_pretrained(MODEL_NAME, from_pt=True)

# model.save_pretrained("./{}".format(MODEL_NAME), saved_model=True)



# import sparknlp
# # let's start Spark with Spark NLP
# spark = sparknlp.start()

# from sparknlp.annotator import *

# xlnet = XlnetEmbeddings.loadSavedModel(
#      '{}/saved_model/1'.format(MODEL_NAME),
#      spark
#  )\
#  .setInputCols(["sentence",'token'])\
#  .setOutputCol("embeddings")\
#  .setCaseSensitive(True)\
#  .setDimension(768)\
#  .setStorageRef('xlnet_base_cased') 

# # pip install transformers --upgrade
# from transformers import AutoTokenizer, XLNetForTokenClassification
# import torch

# tokenizer = AutoTokenizer.from_pretrained("xlnet-base-cased")
# model = XLNetForTokenClassification.from_pretrained('xlnet-base-cased',cache_dir='./cache')

# inputs = tokenizer(
#     "My name is Mrunal", add_special_tokens=False, return_tensors="pt"
# )

# with torch.no_grad():
#     logits = model(**inputs).logits

# predicted_token_class_ids = logits.argmax(-1)

# # Note that tokens are classified rather then input words which means that
# # there might be more predicted token classes than words.
# # Multiple token classes might account for the same word
# predicted_tokens_classes = [model.config.id2label[t.item()] for t in predicted_token_class_ids[0]]

# labels = predicted_token_class_ids
# loss = model(**inputs, labels=labels).loss
# labels



# from transformers import AutoTokenizer, XLNetForTokenClassification
# import torch

# tokenizer = AutoTokenizer.from_pretrained("xlnet-base-cased")
# model = XLNetForTokenClassification.from_pretrained('xlnet-base-cased', cache_dir='./cache')

# inputs = tokenizer(
#     "My name is Mrunal", add_special_tokens=False, return_tensors="pt"
# )

# with torch.no_grad():
#     logits = model(**inputs).logits

# predicted_token_class_ids = logits.argmax(-1)

# # Note that tokens are classified rather than input words, which means that
# # there might be more predicted token classes than words.
# # Multiple token classes might account for the same word
# predicted_tokens_classes = [model.config.id2label[t.item()] for t in predicted_token_class_ids[0]]

# # Create a mapping for class labels to integer IDs
# unique_labels = list(set(predicted_tokens_classes))
# label_to_id = {label: i for i, label in enumerate(unique_labels)}

# # Convert class labels to integer IDs
# label_ids = [label_to_id[label] for label in predicted_tokens_classes]

# # Now, label_ids contains the integer IDs corresponding to the predicted class labels
# print(label_ids)


# from transformers import AutoTokenizer, XLNetForTokenClassification
# import torch

# # Load tokenizer and model
# tokenizer = AutoTokenizer.from_pretrained("xlnet-base-cased")
# model = XLNetForTokenClassification.from_pretrained('xlnet-base-cased', cache_dir='./cache')

# # Define your input text
# input_text = "My name is Mrunal and I work at Google in California."

# # Tokenize the input text
# tokens = tokenizer(input_text, add_special_tokens=False, return_tensors="pt")

# # Make predictions using the model
# with torch.no_grad():
#     logits = model(**tokens).logits

# # Get the predicted token class IDs
# predicted_token_class_ids = logits.argmax(-1)[0].tolist()

# # Get the corresponding token labels using the model's configuration
# predicted_labels = [model.config.id2label[class_id] for class_id in predicted_token_class_ids]

# # Convert the tokenized input text into a list of tokens
# tokens = tokenizer.convert_ids_to_tokens(tokens['input_ids'][0].tolist())

# # Combine consecutive tokens with the same label into named entities
# named_entities = []
# current_entity = {"text": "", "label": None}
# for token, label in zip(tokens, predicted_labels):
#     if label != 'O':
#         if current_entity["label"] is None:
#             current_entity["label"] = label
#         current_entity["text"] += token.replace('▁', ' ')
#     else:
#         if current_entity["label"] is not None:
#             named_entities.append(current_entity)
#             current_entity = {"text": "", "label": None}

# # Print the recognized named entities
# for entity in named_entities:
#     print(f"Entity: {entity['text']}, Label: {entity['label']}")

# from transformers import AutoTokenizer, XLNetForTokenClassification
# import torch

# # Load tokenizer and model
# tokenizer = AutoTokenizer.from_pretrained("xlnet-base-cased")
# model = XLNetForTokenClassification.from_pretrained('xlnet-base-cased', cache_dir='./cache')

# # Define your input text
# input_text = "My name is Mrunal and I work at Google in California."

# # Tokenize the input text
# tokens = tokenizer(input_text, add_special_tokens=False, return_tensors="pt")

# # Make predictions using the model
# with torch.no_grad():
#     logits = model(**tokens).logits

# # Get the predicted token class IDs
# predicted_token_class_ids = logits.argmax(-1)[0].tolist()

# # Get the corresponding token labels using the model's configuration
# predicted_labels = [model.config.id2label[class_id] for class_id in predicted_token_class_ids]

# # Print tokenization and predictions for debugging
# print("Tokens:", tokens)
# print("Predicted Token Class IDs:", predicted_token_class_ids)
# print("Predicted Labels:", predicted_labels)

# # Post-processing to extract named entities
# named_entities = []
# current_entity = {"text": "", "label": None}
# for token, label in zip(tokens, predicted_labels):
#     if label != 'O':
#         if current_entity["label"] is None:
#             current_entity["label"] = label
#         current_entity["text"] += token.replace('▁', ' ')
#     else:
#         if current_entity["label"] is not None:
#             named_entities.append(current_entity)
#             current_entity = {"text": "", "label": None}

# # Print the recognized named entities for debugging
# print("Named Entities:", named_entities)


# from transformers import AutoModelForTokenClassification, AutoTokenizer
# import torch

# # Load tokenizer and NER model
# tokenizer = AutoTokenizer.from_pretrained("dbmdz/bert-large-cased-finetuned-conll03-english")
# model = AutoModelForTokenClassification.from_pretrained("dbmdz/bert-large-cased-finetuned-conll03-english")

# # Define your input text
# input_text = "My name is John and I work at Google in California."

# # Tokenize the input text
# tokens = tokenizer(input_text, add_special_tokens=False, return_tensors="pt")

# # Make predictions using the NER model
# with torch.no_grad():
#     logits = model(**tokens).logits

# # Get the predicted token class IDs
# predicted_token_class_ids = logits.argmax(-1)[0].tolist()

# # Get the corresponding token labels using the model's configuration
# predicted_labels = [model.config.id2label[class_id] for class_id in predicted_token_class_ids]

# # Print the recognized named entities
# named_entities = []
# current_entity = {"text": "", "label": None}
# for token, label in zip(tokens, predicted_labels):
#     if label != 'O':
#         if current_entity["label"] is None:
#             current_entity["label"] = label
#         current_entity["text"] += token.replace('▁', ' ')
#     else:
#         if current_entity["label"] is not None:
#             named_entities.append(current_entity)
#             current_entity = {"text": "", "label": None}

# # Print the recognized named entities
# for entity in named_entities:
#     print(f"Entity: {entity['text']}, Label: {entity['label']}")


# from transformers import pipeline

# # create pipeline for NER
# ner = pipeline('ner', aggregation_strategy = 'simple')


# ner("Sakshi Sunil Patil 703007244 Pillai College of engineering Navi Mumbai")


# tokenizer = AutoTokenizer.from_pretrained("dslim/bert-base-NER")
# model = AutoModelForTokenClassification.from_pretrained("dslim/bert-base-NER")

# nlp = pipeline("ner", model=model, tokenizer=tokenizer)

# example = "John smith 703007244 Pillai College of engineering Navi Mumbai"

# ner_results = nlp(example)


