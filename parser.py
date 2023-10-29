# importing all the libraries
import io
import os
import re
import pandas as pd
import nltk
from nltk.tokenize import sent_tokenize, word_tokenize, TreebankWordTokenizer, WordPunctTokenizer, MWETokenizer
# import PyPDF2
# for pdf to txt
import pdfminer
from pdfminer.pdfinterp import PDFResourceManager, PDFPageInterpreter
from pdfminer.layout import LAParams
from pdfminer.converter import TextConverter
from io import StringIO
from pdfminer.pdfpage import PDFPage
import openpyxl as px
import docx2txt
import constants as cs
import string
# import utils
import pprint
from spacy.matcher import matcher
import multiprocessing as mp
import warnings
from nltk.corpus import stopwords
from nltk.stem import PorterStemmer
from nltk.tokenize import word_tokenize

# from constants import STOPWORDS
warnings.filterwarnings('ignore')

titles = ["Name", "Email", "Phone", "Education", "Experience", "Skills"]
resume_data={}

resume_file=input('Enter resume: ')
print(resume_file)

print(resume_file)


# Extract txt from pdf
# pdf_path="code\Resume.pdf"
def pdf_to_txt(pdf_path):
    resource_manager = PDFResourceManager(caching=True)
    
    # create a string object that will contain the final text the representation of the pdf. 
    out_text = StringIO()
    codec = 'utf-8'
    laParams = LAParams()
    
    # Create a TextConverter Object:
    text_converter = TextConverter(resource_manager, out_text, laparams=laParams)
    fp = open(pdf_path, 'rb')
    #Create a PDF interpreter object

    interpreter = PDFPageInterpreter(resource_manager, text_converter)
    
    # We are going to process the content of each page of the original PDF File    
    for page in PDFPage.get_pages(fp, pagenos=set(), maxpages=0, password="", caching=True, check_extractable=True):
        interpreter.process_page(page)

    
    # Retrieve the entire contents of the “file” at any time 
    text = out_text.getvalue()

    # Closing all the ressources we previously opened

    fp.close()
    text_converter.close()
    out_text.close()
    
    return text
# retext = pdf_to_txt(pdf_path)
# print(pdf_to_txt(pdf_path))


# Extract txt from
# docx_path="Resume.docx"
def docx_to_txt(docx_path):
    temp = docx2txt.process(docx_path)
    text = [line.replace('\t', ' ') for line in temp.split('\n') if line]
    return ' '.join(text)
# print(docx_to_txt(docx_path))


split_tup = os.path.splitext(resume_file)
print(split_tup)
  
# extract the file name and extension
file_name = split_tup[0]
file_extension = split_tup[1]
  
print("File Name: ", file_name)
print("File Extension: ", file_extension)



# detect file extension and call above functions accordingly
def extract_text(file_path, extension):
    '''
    Wrapper function to detect the file extension and call text extraction function accordingly

    :param file_path: path of file of which text is to be extracted
    :param extension: extension of file `file_name`
    '''
    text = ''
    if extension == '.pdf':
        for page in pdf_to_txt(file_path):
            text += '' + page
    elif extension == '.docx' or extension == '.doc':
        text = docx_to_txt(file_path)
    return text
Text= extract_text(resume_file,file_extension)
print(Text)


# Tokenization Of the Extracted Text



nltk.download('all')


# Word Tokenize

#word tokenize
print("word_tokenize",word_tokenize(Text))

tokenizer = TreebankWordTokenizer()
print("TreebankWordTokenizer",tokenizer.tokenize(Text))

tokenizer = WordPunctTokenizer()
print("WordPunctTokenizer",tokenizer.tokenize(Text) )

from nltk.tokenize import MWETokenizer
mtokenizer = MWETokenizer([('athale20comp@studen', 't.mes.ac.in'), ('+91', '-', '8605131403')],separator='')
aftertoken = mtokenizer.tokenize(Text.split())
print("MWETokenizer",aftertoken)
aftertoken=(' '.join(aftertoken))
print(aftertoken)


# Sentence Tokenize



from nltk.tokenize import sent_tokenize
print(sent_tokenize(Text))


# Filteration



def  filter_text(Text,lowerFlag=False,upperFlag=False,numberFlag=False,htmlFlag=False,urlFlag=False,punctFlag=False,spaceFlag=False,charFlag=False,bulletFlag=False,hashtagFlag=False,emojiFlag=False):
    if lowerFlag:
      Text = Text.lower()

    if upperFlag:
      Text = Text.upper()

    if numberFlag:
      import re
      Text = re.sub(r"\d+", '', Text)

    if htmlFlag:
      import re
      Text = re.sub(r'<[^>]*>', '', Text)

    if urlFlag:
      import re
      Text = re.sub(r'(https?|ftp|www)\S+', '', Text)

    if punctFlag:
      import re
      import string
      exclist = string.punctuation #removes [!”#$%&’()*+,-./:;<=>?@[\]^_`{|}~]
      # remove punctuations and digits from oldtext
      table_ = Text.maketrans('', '', exclist)
      Text = Text.translate(table_)

    if spaceFlag:
      import re
      Text = re.sub(' +'," ",Text).strip()

    if hashtagFlag:
      import re
      Text = re.sub(' +',"#",Text).strip()
      
    if charFlag:
      import re
      Text = re.sub('\n', ' ', Text)
      
    if bulletFlag:
      import re
      Text = re.sub('●','',Text)
      Text = re.sub('★','',Text)
      Text = re.sub('|','',Text)
      # Text = re.sub(([0-9|A-Z]+[\.|\)]|)\s+,'', Text)
    
    if emojiFlag:
      import emoji
      Text = emoji.sub(' +',"#",Text).strip()
      pass

    return Text




filteredtxt = filter_text(aftertoken, bulletFlag=True, charFlag=True)
print(filteredtxt)


# Stopwords



# from nltk.corpus import stopwords
stop_words = set(stopwords.words("english"))
print(stop_words)




mtokenizer = MWETokenizer([('athale20comp@studen', 't.mes.ac.in'), ('+91', '-', '8605131403')],separator='')
word_tokens = mtokenizer.tokenize(filteredtxt.split())

# print(ex_text)

stop = [w for w in stop_words if w in word_tokens]
print("StopWords recognized in the given sentence:", stop,"\n")

filtered_sentence = [w for w in word_tokens if not w in stop_words]
filtered_sentence = []

for w in word_tokens:
    if w not in stop_words:
        filtered_sentence.append(w)

print("After removing the recognized stopwords, the Tokens of sentence is:", filtered_sentence)
filtered_data=(' '.join(filtered_sentence))
print(filtered_data)


# Stemming or Lemmatization



import spacy
nlp = spacy.load('en_core_web_sm')

# Create a Doc object
doc = nlp(filtered_data)

# Create list of tokens from given string
tokens = []
for token in doc:
	tokens.append(token)

print(tokens)
#> [the, bats, saw, the, cats, with, best, stripes, hanging, upside, down, by, their, feet]

lemmatized_sentence = " ".join([token.lemma_ for token in doc])

print(lemmatized_sentence)
#> the bat see the cat with good stripe hang upside down by -PRON- foot





def extract_entity_sections(text):
    '''
    Helper function to extract all the raw text from sections of resume

    :param text: Raw text of resume
    :return: dictionary of entities
    '''
    RESUME_SECTIONS = [
                    'accomplishments',
                    'achievements',
                    'experience',
                    'education',
                    'interests',
                    'hobbies',
                    'projects',
                    'professional experience',
                    'experience',
                    'publications',
                    'skills',
                ]
    text_split = [i.strip() for i in text.split('\n')]
    # sections_in_resume = [i for i in text_split if i.lower() in sections]
    entities = {}
    key = False
    for phrase in text_split:
        if len(phrase) == 1:
            p_key = phrase
        else:
            p_key = set(phrase.lower().split()) & set(cs.RESUME_SECTIONS)
        try:
            p_key = list(p_key)[0]
        except IndexError:
            pass
        if p_key in RESUME_SECTIONS:
            entities[p_key] = []
            key = p_key
        elif key and phrase.strip():
            entities[key].append(phrase)
    
    # entity_key = False
    # for entity in entities.keys():
    #     sub_entities = {}
    #     for entry in entities[entity]:
    #         if u'\u2022' not in entry:
    #             sub_entities[entry] = []
    #             entity_key = entry
    #         elif entity_key:
    #             sub_entities[entity_key].append(entry)
    #     entities[entity] = sub_entities

    # # pprint.pprint(entities)

    # make entities that are not found None
    # for entity in cs.RESUME_SECTIONS:
    #     if entity not in entities.keys():
    #         entities[entity] = None 
    
    return entities
pprint.pprint(extract_entity_sections(Text))




def extract_email(text):
    '''
    Helper function to extract email id from text

    :param text: plain text extracted from resume file
    '''
    email = re.findall("([^@|\s]+@[^@]+\.[^@|\s]+)", text)
    if email:
        try:
            return email[0].split()[0].strip(';')
        except IndexError:
            return None
print("E-mail: ",extract_email(filtered_data))
resume_data[titles[1]]=extract_email(filtered_data)




import spacy

# english_nlp = spacy.load('en_core_web_sm')

# text = '''
# This is a sample text that contains the name Alex Smith who is one of the developers of this project.
# You can also find the surname Jones here.
# '''

# spacy_parser = english_nlp(filtered_data)

# for entity in new_nlp.ents:
#     print(f'Found: {entity.text} of type: {entity.label_}')
    
import spacy
from spacy.matcher import Matcher

nlp = spacy.load('en_core_web_sm')
new_matcher = Matcher(nlp.vocab)
new_nlp = nlp(filtered_data)

# def extract_name(text, matcher):
#     '''
#     Helper function to extract name from spacy nlp text

#     :param nlp_text: object of `spacy.tokens.doc.Doc`
#     :param matcher: object of `spacy.matcher.Matcher`
#     :return: string of full name
#     '''
pattern = [{'POS': 'PROPN'}, {'POS': 'PROPN'}, {'POS': 'PROPN'}]
new_matcher.add('NAME',[pattern], on_match=None)
matches = new_matcher(new_nlp)
    
for match_id, start, end in matches:
    span = new_nlp[start:end]
    print("Name: ",span)

# print(extract_name(new_nlp, new_matcher))





#new added code for NER
import pandas as pd
from transformers import XLNetTokenizer, XLNetForTokenClassification, pipeline

# Load your labeled résumé dataset (assuming it's in CSV format)
#resume_data = filteredtxt

# Load the XLNet tokenizer and model
tokenizer = XLNetTokenizer.from_pretrained("xlnet-base-cased")
model = XLNetForTokenClassification.from_pretrained("xlnet-base-cased")

# Create a NER pipeline
nlp_ner = pipeline("ner", model=model, tokenizer=tokenizer)


    # Perform NER
ner_results = nlp_ner(Text)

    # Print the results for this résumé
#print(f"Résumé {index + 1} NER Results:")
for result in ner_results:
    print(f"Entity: {result['word']}, Label: {result['entity']}")




#NER using nltk
import nltk
from nltk import word_tokenize, pos_tag, ne_chunk

# Sample text
#text = "Barack Obama was born in Hawaii and works in Washington, D.C."

# Tokenize the text
tokens = word_tokenize(Text)

# Perform Part-of-Speech tagging
pos_tags = pos_tag(tokens)

# Perform NER using NLTK's ne_chunk
ner_results = ne_chunk(pos_tags)

# Extract named entities
named_entities = []
for subtree in ner_results:
    if type(subtree) == nltk.Tree:
        entity_label = subtree.label()
        entity_text = " ".join([word for word, tag in subtree.leaves()])
        named_entities.append((entity_text, entity_label))

# Print the named entities
for entity, label in named_entities:
    print(f"Entity: {entity}, Label: {label}")




pip install transformers torch





import torch
from transformers import XLNetForSequenceClassification, XLNetTokenizer, AdamW
from torch.utils.data import DataLoader, Dataset
import json

# Load your JSON dataset
with open("traindata3.json", "r", encoding="utf-8") as json_file:
    dataset = json.load(json_file)

# Step 2: Tokenize and Preprocess the Dataset

model_name = "xlnet-base-cased"
tokenizer = XLNetTokenizer.from_pretrained(model_name)

class CustomDataset(Dataset):
    def __init__(self, dataset, tokenizer, max_length=128):
        self.dataset = dataset
        self.tokenizer = tokenizer
        self.max_length = max_length

    def __len__(self):
        return len(self.dataset)

    def __getitem__(self, idx):
        content = self.dataset[idx]['content']
        label = self.dataset[idx]['label']  # Use 'label' as the key for the label field.

        # Tokenize and preprocess the content
        inputs = self.tokenizer(
            content,
            truncation=True,
            padding='max_length',
            max_length=self.max_length,
            return_tensors='pt'
        )

        return {
            'input_ids': inputs['input_ids'].squeeze(),
            'attention_mask': inputs['attention_mask'].squeeze(),
            'label': torch.tensor(label, dtype=torch.long)
        }


# Create instances of CustomDataset for training and validation
train_dataset = CustomDataset(dataset, tokenizer)  # Assuming 'train' is the key for training data.
#val_dataset = CustomDataset(dataset['val'], tokenizer)      # Assuming 'val' is the key for validation data.

# Step 3: Fine-Tuning Configuration

learning_rate = 2e-5
num_epochs = 3
batch_size = 32

# Step 4: Load Pretrained XLNet Model
model = XLNetForSequenceClassification.from_pretrained(model_name, num_labels=2)  # Change num_labels based on your task.

# Step 5: Prepare Data Loaders
train_dataloader = DataLoader(train_dataset, batch_size=batch_size, shuffle=True)
#val_dataloader = DataLoader(val_dataset, batch_size=batch_size)




for batch in train_dataloader:
    print(batch.keys())




dataset[0]




# Step 6: Define Training Loop
optimizer = AdamW(model.parameters(), lr=learning_rate)

for epoch in range(num_epochs):
    model.train()
    for batch in train_dataloader:
        inputs = batch['input_ids'].to(device)
        attention_mask = batch['attention_mask'].to(device)
        labels = batch['label'].to(device)

        outputs = model(inputs, attention_mask=attention_mask, labels=labels)
        loss = outputs.loss

        loss.backward()
        optimizer.step()
        optimizer.zero_grad()

# Step 7: Evaluate Model
model.eval()
with torch.no_grad():
    correct = 0
    total = 0

    for batch in val_dataloader:
        inputs = batch['input_ids'].to(device)
        attention_mask = batch['attention_mask'].to(device)
        labels = batch['label'].to(device)

        outputs = model(inputs, attention_mask=attention_mask)
        logits = outputs.logits

        predicted_labels = torch.argmax(logits, dim=1)
        correct += (predicted_labels == labels).sum().item()
        total += len(labels)

    accuracy = correct / total
    print(f"Validation Accuracy: {accuracy}")

# Step 8: Save the Fine-Tuned Model
model.save_pretrained("fine_tuned_xlnet")




# Specify the input and output file paths
input_file_path = "traindata.json"
output_file_path = "traindata3.json"

# Open the input and output files
with open(input_file_path, "r", encoding="utf-8") as input_file, \
     open(output_file_path, "w", encoding="utf-8") as output_file:

    # Initialize a flag to track if it's the first line
    first_line = True

    # Iterate through the lines in the input file
    for line in input_file:
        # Add a comma before each line (except the first line)
        if not first_line:
            output_file.write(",")
        else:
            first_line = False

        # Write the line from the input file to the output file
        output_file.write(line)

    # Write a closing square bracket to complete the JSON array
    output_file.write("]")





def extract_mobile_number(text):
    '''
    Helper function to extract mobile number from text

    :param text: plain text extracted from resume file
    :return: string of extracted mobile numbers
    '''
    # Found this complicated regex on : https://zapier.com/blog/extract-links-email-phone-regex/
    phone = re.findall(re.compile(r'(?:(?:\+?([1-9]|[0-9][0-9]|[0-9][0-9][0-9])\s*(?:[.-]\s*)?)?(?:\(\s*([2-9]1[02-9]|[2-9][02-8]1|[2-9][02-8][02-9])\s*\)|([0-9][1-9]|[0-9]1[02-9]|[2-9][02-8]1|[2-9][02-8][02-9]))\s*(?:[.-]\s*)?)?([2-9]1[02-9]|[2-9][02-9]1|[2-9][02-9]{2})\s*(?:[.-]\s*)?([0-9]{4})(?:\s*(?:#|x\.?|ext\.?|extension)\s*(\d+))?'), text)
    if phone:
        number = ''.join(phone[0])
        if len(number) > 10:
            return '+' + number
        else:
            return number
print("Mobile No.: ",extract_mobile_number(filtered_data))
resume_data[titles[2]]=extract_mobile_number(filtered_data)




def list_to_string(lst):
    # Convert the list to a string representation
    return ', '.join(map(str, lst))
noun_chunks=[]
extracted_skills=[]
file='skills.csv'
new_nlp = nlp(filtered_data)
def extract_skills(nlp_text, noun_chunks):
    '''
    Helper function to extract skills from spacy nlp text

    :param nlp_text: object of `spacy.tokens.doc.Doc`
    :param noun_chunks: noun chunks extracted from nlp text
    :return: list of skills extracted
    '''
    tokens = [token.text for token in nlp_text if not token.is_stop]
    data = pd.read_csv(os.path.join(os.path.dirname(file), 'skills.csv')) 
    skills = list(data.columns.values)
    skillset = []
    # check for one-grams
    for token in tokens:
        if token.lower() in skills:
            skillset.append(token)
    
    # check for bi-grams and tri-grams
    for token in noun_chunks:
        token = token.text.lower().strip()
        if token in skills:
            skillset.append(token)
    extracted_skills=[i.capitalize() for i in set([i.lower() for i in skillset])]

    return [i.capitalize() for i in set([i.lower() for i in skillset])]
extract_skills(new_nlp,noun_chunks)
extracted_skills=extract_skills(new_nlp,noun_chunks)
print(extracted_skills)
resume_data[titles[5]]=list_to_string(extracted_skills)



import openpyxl
from openpyxl import Workbook
from datetime import datetime
import pandas as pd

def create_excel_sheet(data):
    # Create a new Excel workbook
    wb = Workbook()
    
    # Get the current timestamp for the sheet name
    timestamp = datetime.now().strftime("%Y-%m-%d %H-%M-%S")
    
    # Create a new sheet with the timestamp as the name
    sheet = wb.create_sheet(title=timestamp)
    
    # Create a DataFrame from the list
    df = pd.DataFrame.from_dict(resume_data, orient='index', columns=['Text'])


    
    # Write the data to the Excel sheet
    for index, row in df.iterrows():
        sheet.append(row.tolist())
    
    # Save the Excel workbook
    wb.save(f"output_{timestamp}.xlsx")

if __name__ == "__main__":
    # Sample list to convert
    my_list = extract_skills(new_nlp,noun_chunks)
    
    # Convert the list to a single string
    converted_data = list_to_string(my_list)
    
    # Create a new Excel sheet with the converted data
    create_excel_sheet([converted_data])





import openpyxl
from openpyxl import Workbook
from datetime import datetime
import pandas as pd

def create_excel_sheet(dataframe):
    # Create a new Excel workbook
    wb = Workbook()
    
    # Get the current timestamp for the sheet name
    timestamp = datetime.now().strftime("%Y-%m-%d %H-%M-%S")
    
    # Create a new sheet with the timestamp as the name
    sheet = wb.create_sheet(title=timestamp)
    
    # Convert the DataFrame to a list of lists for easy insertion
    data = dataframe.values.tolist()
    
    # Write the data to the Excel sheet
    for row in data:
        sheet.append(row)
    
    # Save the Excel workbook
    wb.save(f"output_{timestamp}.xlsx")

if __name__ == "__main__":
    # Sample DataFrame
    data = {'Column1': [1, 2, 3, 4],
            'Column2': ['A', 'B', 'C', 'D']}
    df = pd.DataFrame(data)
    
    # Create a new Excel sheet with the DataFrame
    create_excel_sheet(df)




print(list_to_string(my_list))


# Matching the extracted skills with our requirements



file='required.csv'
new_nlp = nlp(filtered_data)
from difflib import SequenceMatcher

def match_skills(skills):
    '''
    Helper function to extract skills from spacy nlp text

    :param nlp_text: object of `spacy.tokens.doc.Doc`
    :param noun_chunks: noun chunks extracted from nlp text
    :return: list of skills extracted
    '''
    # tokens = [token.text for token in nlp_text if not token.is_stop]
    # skills=[]
    data = pd.read_csv(os.path.join(os.path.dirname(file), 'required.csv')) 
    required = list(data.columns.values)
    print("Skills in Resume: ",skills)
    print("Required skills are: ",required)
    matched = []
    # check for one-grams
    for skill in skills:
        if skill.lower() in required:
            matched.append(skill)
    print("Matched skills are: ",matched)
    res = len(set(required) and set(matched)) / float(len(set(required) or set(matched))) * 100
    print("The percentage of skills matched is: ",res)

match_skills(extracted_skills)




def extract_education(nlp_text):
    '''
    Helper function to extract education from spacy nlp text

    :param nlp_text: object of `spacy.tokens.doc.Doc`
    :return: tuple of education degree and year if year if found else only returns education degree
    '''
    EDUCATION         = [
                    'BE','B.E.', 'B.E', 'BS', 'B.S', 'ME', 'M.E', 'M.E.', 'MS', 'M.S', 'BTECH', 'MTECH', 
                    'SSC', 'HSC', 'CBSE', 'ICSE', 'X', 'XII'
                    ]
    
    STOPWORDS         = set(stopwords.words('english'))
    
    YEAR              = r'(((20|19)(\d{2})))'
    
    edu = {}
    # Extract education degree
    for index, text in enumerate(nlp_text):
        for tex in text.split():
            # tex = re.sub(r'[?|$|.|!|,]', r'', tex)
            if tex.upper() in EDUCATION:
                edu[tex] = text + nlp_text[index + 1]

    # Extract year
    education = []
    for key in edu.keys():
        year = re.search(re.compile(YEAR), edu[key])
        if year:
            education.append((key, ''.join(year.group(0))))
        else:
            education.append(key)
    return education
print("Education Qualification: ",extract_education(Text))




from nltk.stem import WordNetLemmatizer 
from nltk.corpus import stopwords

def extract_experience(resume_text):
    '''
    Helper function to extract experience from resume text

    :param resume_text: Plain resume text
    :return: list of experience
    '''
    wordnet_lemmatizer = WordNetLemmatizer()
    stop_words = set(stopwords.words('english'))

    # word tokenization 
    word_tokens = nltk.word_tokenize(resume_text)

    # remove stop words and lemmatize  
    filtered_sentence = [w for w in word_tokens if not w in stop_words and wordnet_lemmatizer.lemmatize(w) not in stop_words] 
    sent = nltk.pos_tag(filtered_sentence)

    # parse regex
    cp = nltk.RegexpParser('P: {<NNP>+}')
    cs = cp.parse(sent)
    
    # for i in cs.subtrees(filter=lambda x: x.label() == 'P'):
    #     print(i)
    
    test = []
    
    for vp in list(cs.subtrees(filter=lambda x: x.label()=='P')):
        test.append(" ".join([i[0] for i in vp.leaves() if len(vp.leaves()) >= 2]))

    # Search the word 'experience' in the chunk and then print out the text after it
    x = [x[x.lower().index('experience') + 10:] for i, x in enumerate(test) if x and 'experience' in x.lower()]
    return x
print(extract_experience(Text))




exp_list=[]

def extract_competencies(text, experience_list):
    '''
    Helper function to extract competencies from resume text
    :param resume_text: Plain resume text
    :return: dictionary of competencies
    '''
    experience_text = ' '.join(experience_list)
    competency_dict = {}
    COMPETENCIES = {
    'teamwork': [
        'supervised',
        'facilitated',
        'planned',
        'plan',
        'served',
        'serve',
        'project lead',
        'managing',
        'managed',
        'lead ',
        'project team',
        'team',
        'conducted',
        'worked',
        'gathered',
        'organized',
        'mentored',
        'assist',
        'review',
        'help',
        'involve',
        'share',
        'support',
        'coordinate',
        'cooperate',
        'contributed'
    ],
    'communication': [
        'addressed',
        'collaborated',
        'conveyed',
        'enlivened',
        'instructed',
        'performed',
        'presented',
        'spoke',
        'trained',
        'author',
        'communicate',
        'define',
        'influence',
        'negotiated',
        'outline',
        'proposed',
        'persuaded',
        'edit',
        'interviewed',
        'summarize',
        'translate',
        'write',
        'wrote',
        'project plan',
        'business case',
        'proposal',
        'writeup'
    ],
    'analytical': [
        'process improvement',
        'competitive analysis',
        'aligned',
        'strategive planning',
        'cost savings',
        'researched ',
        'identified',
        'created',
        'led',
        'measure',
        'program',
        'quantify',
        'forecasr',
        'estimate',
        'analyzed',
        'survey',
        'reduced',
        'cut cost',
        'conserved',
        'budget',
        'balanced',
        'allocate',
        'adjust',
        'lauched',
        'hired',
        'spedup',
        'speedup',
        'ran',
        'run',
        'enchanced',
        'developed'
    ],
    'result_driven': [
        'cut',
        'decrease',
        'eliminate',
        'increase',
        'lower',
        'maximize',
        'rasie',
        'reduce',
        'accelerate',
        'accomplish',
        'advance',
        'boost',
        'change',
        'improve',
        'saved',
        'save',
        'solve',
        'solved',
        'upgrade',
        'fix',
        'fixed',
        'correct',
        'achieve'           
    ],
    'leadership': [
        'advise',
        'coach',
        'guide',
        'influence',
        'inspire',
        'instruct',
        'teach',
        'authorized',
        'chair',
        'control',
        'establish',
        'execute',
        'hire',
        'multi-task',
        'oversee',
        'navigate',
        'prioritize',
        'approve',
        'administer',
        'preside',
        'enforce',
        'delegate',
        'coordinate',
        'streamlined',
        'produce',
        'review',
        'supervise',
        'terminate',
        'found',
        'set up',
        'spearhead',
        'originate',
        'innovate',
        'implement',
        'design',
        'launch',
        'pioneer',
        'institute'
    ]
}
    for competency in COMPETENCIES.keys():
        for item in COMPETENCIES[competency]:
            if string_found(item, experience_text):
                if competency not in competency_dict.keys():
                    competency_dict[competency] = [item]
                else:
                    competency_dict[competency].append(item)
    return competency_dict
print(extract_competencies(Text, exp_list))    



def extract_measurable_results(text, experience_list):
    '''
    Helper function to extract measurable results from resume text

    :param resume_text: Plain resume text
    :return: dictionary of measurable results
    '''

    # we scan for measurable results only in first half of each sentence
    experience_text = ' '.join([text[:len(text) // 2 - 1] for text in experience_list])
    mr_dict = {}

    for mr in cs.MEASURABLE_RESULTS.keys():
        for item in cs.MEASURABLE_RESULTS[mr]:
            if string_found(item, experience_text):
                if mr not in mr_dict.keys():
                    mr_dict[mr] = [item]
                else:
                    mr_dict[mr].append(item)
    
    return mr_dict



# ner using xlnet testing 



import torch
from transformers import XLNetForSequenceClassification, XLNetTokenizer, AdamW
from torch.utils.data import DataLoader, Dataset
import json

# Load your JSON dataset
with open("traindata3.json", "r", encoding="utf-8") as json_file:
    data = json.load(json_file)




data[0]





def process_label_and_points(label, points):
    # Initialize an empty list to store the extracted entity texts
    entity_texts = []
    
    # Loop through each point dictionary in the list
    for point in points:
        start = point['start']
        end = point['end']
        
        # Extract the labeled entity text from the content
        entity_text = content[start:end]
        
        # Append the entity text to the list
        entity_texts.append(entity_text)
    
    # Combine the extracted entity texts if needed (e.g., join with a space)
    combined_entity_text = ' '.join(entity_texts)
    
    return combined_entity_text





input_data




from transformers import XLNetTokenizer
import torch

# Initialize the XLNet tokenizer
tokenizer = XLNetTokenizer.from_pretrained('xlnet-base-cased')

# Tokenize and pad/truncate input sequences
input_ids = []
attention_masks = []

for text in Text:
    encoded_text = tokenizer.encode_plus(
        text,
        add_special_tokens=True,
        max_length=128,  # Adjust the max length as needed
        padding='max_length',
        return_attention_mask=True,
        return_tensors='pt'
    )

    input_ids.append(encoded_text['input_ids'])
    attention_masks.append(encoded_text['attention_mask'])

# Convert input_ids and attention_masks to tensors
input_ids = torch.cat(input_ids, dim=0)
attention_masks = torch.cat(attention_masks, dim=0)

# Encode the target labels (NER tags) into tensors or other suitable format
# You may need to use label encoders or other methods based on your model's requirements



input_ids




import torch
from torch.utils.data import Dataset

class NERDataset(Dataset):
    def __init__(self, input_ids, attention_masks, target_labels):
        """
        Args:
            input_ids (list of torch.Tensor): List of input sequences as tensors.
            attention_masks (list of torch.Tensor): List of attention masks as tensors.
            target_labels (list of list of str): List of target labels for NER, where each list contains labels for one example.
        """
        self.input_ids = input_ids
        self.attention_masks = attention_masks
        self.target_labels = target_labels

    def __len__(self):
        return len(self.input_ids)

    def __getitem__(self, idx):
        sample = {
            'input_ids': self.input_ids[idx],
            'attention_mask': self.attention_masks[idx],
            'target_labels': self.target_labels[idx],
        }
        return sample




from transformers import XLNetForTokenClassification, Trainer, TrainingArguments
label_list = ['College Name', 'Designation', 'Name', 'Companies worked at', 'Email Address', 'Location', 'Skills']
target_labels = [
    ["O", "O", "B-ORG", "I-ORG", "O", ],  # Labels for the first example
    ["B-PER", "I-PER", "O", "B-LOC", "O", ]  # Labels for the second example
    # Add labels for more examples...
]

# Calculate the number of unique labels
num_labels = len(label_list)# Load the pre-trained XLNet model for token classification
model = XLNetForTokenClassification.from_pretrained('xlnet-base-cased', num_labels=num_labels)  # Define num_labels accordingly

# Define training arguments (batch size, learning rate, etc.)
training_args = TrainingArguments(
    output_dir='./ner_model',
    per_device_train_batch_size=16,
    num_train_epochs=3,
    learning_rate=2e-5,
)

# Initialize the Trainer
trainer = Trainer(
    model=model,
    args=training_args,
    train_dataset=NERDataset(input_ids, attention_masks, target_labels),  # Define your custom dataset class
)

# Start training
trainer.train()




pip install transformers[torch]




pip install accelerate -U










