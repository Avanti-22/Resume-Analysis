from django.shortcuts import get_object_or_404, render, redirect
from django.contrib.auth.decorators import login_required
from django.contrib.auth.models import User
from django.contrib.auth import authenticate, logout, login
from datetime import datetime
from django.contrib import messages
from django.db import models
from webapp.models import Contact, Matched, Resumeform, JobDescription, ResumeData
import os
from .matching import fetch_data, match_skills
from django.db.models import Avg
from django.db.models import F, Value

from .models import ResumeData,Matched
from django.db import models
from django.shortcuts import render
import requests
from openpyxl import load_workbook
from bs4 import BeautifulSoup

#from .models import Details

from .matching import fetch_data

# from .forms import PDFUploadForm

# Create your views here.
#--------------------common views-----------------------------

def home(request):
    return render(request,'home.html')
# @login_required()

def aboutView(request):
    return render(request,'aboutus.html')

def contactView(request):
    if request.method =="POST":
        name = request.POST.get('name')
        email = request.POST.get('email')
        subject = request.POST.get('subject')
        message = request.POST.get('message')
        contact = Contact(Name=name, Email=email, Subject=subject, Message=message, Date=datetime.today())
        contact.save()
        messages.success(request, "Your message has been sent.")
    return render(request,'contactus.html')

def faqView(request):
    return render(request,'faq.html')

def userlogout(request):
    logout(request)
    return render(request,'home.html')

#-------------------------User views----------------------------


def userview(request):
    return render(request, 'candidateview.html')

def userreg(request):
    if request.method == 'POST':
        # name = request.POST['name']
        email = request.POST.get('email')
        username = request.POST.get('username')
        password = request.POST.get('password')
        cpassword = request.POST.get('cpassword')
        # print(email, username, password, cpassword)
        if password == cpassword:
            if User.objects.filter(username=username).exists():
                messages.info(
                    request, 'This username already exists. Please try again!!')
                return render(request, 'register.html')
            else:
                user = User(
                    email=email, username=username, password=password)
                user.set_password(password)
                user.save()
                
                messages.info(
                    request, 'Account created successfully.')
                return render(request,'login.html')
        else:
            messages.info(
                request, 'The Passwords doesnot match. Please try again!!')
            return render(request, 'register.html')

    return render(request, 'register.html')

def userlogin(request):
    if request.method=="POST":
        username=request.POST.get('username')
        password=request.POST.get('password')
        #validate the user
        # # print(username,password)

        user = authenticate(username=username, password=password)
        if user is not None:
            # A backend authenticated the credentials
            login(request, user)
            return render(request,'upload.html')    
        else:
            # No backend authenticated the credentials
            return render(request, 'login.html')

    return render(request, 'login.html')

def upload(request):
    if request.user.is_anonymous:
        return render(request, 'login.html')

    if request.method == 'POST':
        name = request.POST.get('name')
        email = request.POST.get('email')
        resumefile = request.FILES.get('file')  # Use request.FILES to access the uploaded file

        if resumefile:
            # Create the directory if it doesn't exist
            upload_folder = 'pdfs'
            os.makedirs(upload_folder, exist_ok=True)

            # Save the uploaded file to the specified location
            with open(os.path.join(upload_folder, resumefile.name), 'wb') as destination:
                for chunk in resumefile.chunks():
                    destination.write(chunk)

            # Create and save the model instance
            resumeform = Resumeform(R_Name=name, R_Email=email, Resumefile=os.path.join(upload_folder, resumefile.name))
            resumeform.save()
            resume_id = resumeform.Resume_No
            # print(resume_id)
            request.session['resume_id'] = resume_id
            messages.success(request, "Your form has been submitted.")
            jobs = JobDescription.objects.all()
            # print(jobs)# Assuming Job is your model containing job details
        return render(request, 'jd_for_matching.html', {'jobs': jobs})  # Redirect to a success page
    else:
        # Handle GET request or render the initial form
        return render(request, 'upload.html')

    # return render(request, 'upload.html')

def jd_for_matching(request):
    jobs = JobDescription.objects.all()
    # print(jobs)# Assuming Job is your model containing job details
    return render(request, 'jd_for_matching.html', {'jobs': jobs})

def resume_matching(request, job_id):
    resume_id = request.session.get('resume_id')
    # print(resume_id)
    # Retrieve Resumeform and JobDescription objects
    resume = get_object_or_404(Resumeform, Resume_No=resume_id)
    job_description = get_object_or_404(JobDescription, Job_id=job_id)

    # Retrieve required skills from JobDescription
    required_skills = job_description.Required_Skills

    # Check if ResumeData exists for the current resume_id
    try:
        resume_data = ResumeData.objects.get(Resume_id=resume_id)

        # If ResumeData exists, retrieve relevant data
        email = resume_data.Email
        mobile_number = resume_data.Mobile_No
        education = resume_data.Education
        skills = resume_data.Skills
        name =resume_data.Name
        # Calculate match percentage using match_skills function
        # match_percent = match_skills(skills, required_skills)

    except ResumeData.DoesNotExist:
        # If ResumeData does not exist, fetch data using fetch_data function
        resume_content = fetch_data(resume_id, job_id)

        # Extract data from fetch_data result
        email = resume_content[1]
        mobile_number = resume_content[2]
        education = resume_content[3]
        skills = resume_content[4]
        name = resume.R_Name
        # Create new ResumeData object
        resume_data = ResumeData.objects.create(
            Resume_id=resume_id,
            Name= name,
            Email=email,
            Mobile_No=mobile_number,
            Education=education,
            Skills=skills
        )

        # Calculate match percentage using match_skills function
    match_result = match_skills(required_skills, skills)
    matched_skills = match_result[1]
        # Round match percentage to two decimal places
    match_percent = round(match_result[0], 2)

    # Create Matched object to store matching details
    matched_data = Matched.objects.create(
        Extracted_skills=skills,
        Required_skills=required_skills,
        Matched_skills=matched_skills,
        Percent_matched=match_percent,
        Resume_id=resume_id,
        Job_id=job_id,
        Name=name,  # Access name from related Resumeform
        Email=email,
        Mobile_No=mobile_number,
    )

    # Pass data to matching.html template for rendering
    return render(request, 'matching.html', {
        'resume': resume,
        'name':name,
        'email': email,
        'mobile_number': mobile_number,
        'education': education,
        'skills': skills,
        'job_description': job_description,
        'match_percent': match_percent
    })
    
#------------------------HR views----------------------

def hrview(request):
    
    if request.method=="POST":
        username=request.POST.get('username')
        password=request.POST.get('password')
        #validate the user
        # print(username,password)

        user = authenticate(username=username, password=password)
        if user is not None:
            # A backend authenticated the credentials
            login(request, user)
            return render(request,'hr_dash.html')    
        else:
            # No backend authenticated the credentials
            return render(request, 'hrview.html')

    return render(request, 'hrview.html')

def hr_dash(request):
    if request.user.is_superuser:
        # A backend authenticated the credentials
        return render(request,'hr_dash.html')    
    else:
        # No backend authenticated the credentials
        return render(request, 'hrview.html')

def hr_jd(request):
    uploaded_resume_ids=[]
    
    if request.method == 'POST':
        # Extract form data from POST request
        job_title = request.POST.get('job_title')
        job_description = request.POST.get('job_description')
        requirements = request.POST.get('requirements')
        resumefiles = request.FILES.getlist('files')  # Use request.FILES.getlist to access multiple uploaded files
        # experience = request.POST.get('experience')
        
        for resumefile in resumefiles:
                
            if resumefile:
                # Create the directory if it doesn't exist
                upload_folder = 'pdfs'
                os.makedirs(upload_folder, exist_ok=True)
                
                # Process each uploaded file
            
                # Save the uploaded file to the specified location
                with open(os.path.join(upload_folder, resumefile.name), 'wb') as destination:
                    for chunk in resumefile.chunks():
                        destination.write(chunk)

                # Create and save the model instance for each resume
                resumeform = Resumeform(Resumefile=os.path.join(upload_folder, resumefile.name))
                resumeform.save()
                uploaded_resume_id = resumeform.Resume_No
                uploaded_resume_ids.append(uploaded_resume_id)
        
        request.session['uploaded_resume_ids'] = uploaded_resume_ids
        
        job_des = JobDescription(Title=job_title, Required_Skills=requirements)
        job_des.save()
        new_job_id = job_des.Job_id
        request.session['new_job_id'] = new_job_id
        job = get_object_or_404(JobDescription, Job_id=new_job_id)
        #matching function call
        matching_with_new_jd(new_job_id)
        
        messages.success(request, "Your Job Description has been added.")
        
        return render(request, 'after_new_jd.html', {'job': job})
    else:
        return render(request, 'hr_jd.html')

def after_new_jd(request):
    new_job_id= request.session.get('new_job_id')
    uploaded_resume_ids= request.session.get('uploaded_resume_ids')
    job = get_object_or_404(JobDescription, Job_id=new_job_id)
    # # print(job)
    matching_uploaded_with_new_jd(new_job_id, uploaded_resume_ids)
    return render(request, 'after_new_jd.html', {'job': job})
    
def jd_for_ranking(request):
    jobs = JobDescription.objects.all()
    # print(jobs)# Assuming Job is your model containing job details
    return render(request, 'jd_for_ranking.html', {'jobs': jobs})

def batch_resume_ranking(request, job_id):
    resume_ids= request.session.get('uploaded_resume_ids')
    matching_uploaded_with_new_jd(job_id, resume_ids)
    # print("resume ids are:",resume_ids)
    job= get_object_or_404(JobDescription, Job_id=job_id)
    # print(job)
    resumedata_objects = ResumeData.objects.filter(Resume_id__in=resume_ids)
    
    matched_data = (
        Matched.objects
        .filter(Job_id=job_id, Resume_id__in=resume_ids)
        .annotate(
            percent_matched=F('Percent_matched'),
            name=F('Name'),  # Access name from related Resumeform
            email=F('Email'),
            resume_id=F('Resume_id'),  # Access Resume_id from Matched model
            extracted_skills=F('Extracted_skills')
        )
        .values('percent_matched', 'name', 'email', 'resume_id', 'extracted_skills')
        .order_by('-percent_matched')
    )
    # print(matched_data)
    # Convert the queryset to a list of dictionaries
    matched_list = list(matched_data)
    # return matched_data
    # Render template with ranked resumes
    return render(request, 'resume_ranking.html', {'resumes': matched_data, 'job': job})

def resume_ranking(request, job_id):
    
    job= get_object_or_404(JobDescription, Job_id=job_id)
    # print(job)
    # Filter Matched records for the given job_id and annotate required fields
    matched_data = (
        Matched.objects
        .filter(Job_id=job_id)
        .annotate(
            percent_matched=F('Percent_matched'),
            name=F('Resume__R_Name'),  # Access name from related Resumeform
            email=F('Email'),
            mobile=F('Mobile_No'),
            resume_id=F('Resume_id'),  # Access Resume_id from Matched model
            extracted_skills=F('Extracted_skills')
        )
        .values('percent_matched', 'name', 'email', 'mobile', 'resume_id', 'extracted_skills')
        .order_by('-percent_matched')
    )
    # print(matched_data)
    # Convert the queryset to a list of dictionaries
    matched_list = list(matched_data)
    # return matched_data
    # Render template with ranked resumes
    return render(request, 'resume_ranking.html', {'resumes': matched_data, 'job': job})

def import_from_excel(request):
    if request.method == 'POST':
        excel_file = request.FILES[r"ResumeProcessing\Extracted Data\Extracted.xlsx"]
        wb = load_workbook(excel_file)
        ws = wb.active

        for row in ws.iter_rows(min_row=2, values_only=True):
            name, email, phone_number, skills = row
            Details.objects.create(name=name, email=email, phone_number=phone_number, skills=skills)

        return render(request, 'import_success.html')

    return render(request, 'import_form.html')

def display_data(request):
    data = Details.objects.all()
    return render(request, 'display_data.html', {'data': data})


# ----------------------------matching functions -------------------------------
def matching_with_new_jd(job_id):
    job = get_object_or_404(JobDescription, Job_id=job_id)
    # Retrieve all ResumeData objects
    resumedata_objects = ResumeData.objects.all()
    required_skills = job.Required_Skills
    # Process each ResumeData object to match skills with the new JobDescription
    for resume_data in resumedata_objects:
        # resume_data = ResumeData.objects.get(Resume_id=resume_id)

        # If ResumeData exists, retrieve relevant data
        email = resume_data.Email
        mobile_number = resume_data.Mobile_No
        education = resume_data.Education
        skills = resume_data.Skills
        resume_id = resume_data.Resume_id
        # Calculate match percentage using match_skills function
        match_result = match_skills(required_skills, skills)
        matched_skills = match_result[1]
        # Round match percentage to two decimal places
        match_percent = round(match_result[0], 2)


        # Create Matched object to store matching details
        matched_data = Matched.objects.create(
            Extracted_skills=skills,
            Required_skills=required_skills,
            Matched_skills=matched_skills,
            Percent_matched=match_percent,
            Resume_id=resume_id,                
            Job_id=job_id,
            Email=email,
            Mobile_No=mobile_number,
           )

def matching_uploaded_with_new_jd(job_id,uploaded_resume_ids):
    job = get_object_or_404(JobDescription, Job_id=job_id)
    required_skills = job.Required_Skills
    
    uploaded_resumes = Resumeform.objects.filter(Resume_No__in=uploaded_resume_ids)
    
    for uploaded_resume in uploaded_resumes:
        uploaded_resume_id= uploaded_resume.Resume_No
        resume_content = fetch_data(uploaded_resume_id, job_id)
        # Extract data from fetch_data result
        email = resume_content[1]
        mobile_number = resume_content[2]
        education = resume_content[3]
        skills = resume_content[4]

        # Create new ResumeData object
        resume_data = ResumeData.objects.create(
            Resume_id=uploaded_resume_id,
            Email=email,
            Mobile_No=mobile_number,
            Education=education,
            Skills=skills
        )
        
        match_result = match_skills(required_skills, skills)
        matched_skills = match_result[1]
        # Round match percentage to two decimal places
        match_percent = round(match_result[0], 2)
        # Create Matched object to store matching details
        matched_data = Matched.objects.create(
                    Extracted_skills=skills,
                    Required_skills=required_skills,
                    Matched_skills=matched_skills,
                    Percent_matched=match_percent,
                    Resume_id=uploaded_resume_id,                
                    Job_id=job_id,
                    Email=email,
                    Mobile_No=mobile_number,
                    )
        
def remote_jobs_view(request):
    url = 'https://remotive.com/api/remote-jobs'
    response = requests.get(url)
    data = response.json()
    jobs = data.get('jobs', [])

    # Remove HTML tags from job descriptions
    for job in jobs:
        html_content = job.get('description', '')
        # Parse HTML content using BeautifulSoup
        soup = BeautifulSoup(html_content, 'html.parser')
        # Get text without HTML tags
        text_without_html = soup.get_text(separator=' ')
        # Update job description with text without HTML tags
        job['description'] = text_without_html.strip()

    return render(request, 'jobs.html', {'jobs': jobs})
