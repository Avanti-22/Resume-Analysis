# importing all the libraries
import io
import os
import re
from django.shortcuts import get_object_or_404
import pandas as pd
import nltk
from nltk.tokenize import sent_tokenize, word_tokenize, TreebankWordTokenizer, WordPunctTokenizer, MWETokenizer
# import PyPDF2
# for pdf to txt
import pdfminer
from pdfminer.pdfinterp import PDFResourceManager, PDFPageInterpreter
from pdfminer.layout import LAParams
from pdfminer.converter import TextConverter
from io import StringIO
from pdfminer.pdfpage import PDFPage
import openpyxl as px
import docx2txt
import constants as cs
import string
# import utils
import pprint
# from sklearn import pipeline
from spacy.matcher import matcher
import multiprocessing as mp
import warnings
from nltk.corpus import stopwords
from nltk.stem import PorterStemmer
from nltk.tokenize import word_tokenize
from django.core.files.storage import default_storage
from webapp.models import JobDescription, Resumeform
from openpyxl import Workbook
from datetime import datetime
import pandas as pd
from math import sqrt
import tensorflow as tf
from transformers import AutoTokenizer, AutoModelForTokenClassification
from transformers import pipeline
# from constants import STOPWORDS
warnings.filterwarnings('ignore')
filtered_data=[]
resume_data=[]
titles = ["Name", "Email", "Phone", "Education", "Experience", "Skills"]
resume_data={}
text={}
# main function
def fetch_data(resume_id, job_id):
    
    # Retrieve the resume object using the ID
    resume = get_object_or_404(Resumeform, Resume_No=resume_id)
    job = get_object_or_404(JobDescription, Job_id=job_id)
    
    resume_file = resume.Resumefile
    Text=extract_text(resume_file)

    # Tokenization Of the Extracted Text
    tokenizer = WordPunctTokenizer()
    mtokenizer = MWETokenizer([('athale20comp@studen', 't.mes.ac.in'), ('+91', '-', '8605131403')], separator='')
    aftertoken = mtokenizer.tokenize(Text.split())
    aftertoken = ' '.join(aftertoken)

    filteredtxt = filter_text(aftertoken, bulletFlag=True, charFlag=True)

    # Stopword removal
    stop_words = set(stopwords.words("english"))
    word_tokens = mtokenizer.tokenize(filteredtxt.split())

    filtered_sentence = [w for w in word_tokens if not w in stop_words]
    filtered_data = ' '.join(filtered_sentence)
    # # print(filtered_data)

    
    resume_data[titles[1]]= extract_name(filtered_data)
    # Extract Email
    # # print("E-mail: ", extract_email(filtered_data))
    resume_data[titles[2]] = extract_email(filtered_data)

    # Stemming or Lemmatization
    import spacy

    nlp = spacy.load('en_core_web_sm')
    doc = nlp(filtered_data)
    lemmatized_sentence = " ".join([token.lemma_ for token in doc])
    # # print(lemmatized_sentence)

    # Extract Name
    from spacy.matcher import Matcher
    new_matcher = Matcher(nlp.vocab)
    new_nlp = nlp(filtered_data)
    # # print(extract_name(new_nlp, new_matcher))

    # Extract Mobile Number
    # # print("Mobile No.: ", extract_mobile_number(filtered_data))
    resume_data[titles[3]] = extract_mobile_number(filtered_data)

    # Extract Education
    Education_Qualification = extract_education(filteredtxt)
    # # print("Education_Qualification: ", Education_Qualification)
    resume_data[titles[4]] = list_to_string(Education_Qualification)
    
    # Extract Skills
    noun_chunks = []
    extracted_skills=[]
    new_nlp = nlp(filtered_data)
    
    extracted_skills = extract_skills(new_nlp, noun_chunks)
    # # print(extracted_skills)
    resume_data[titles[5]] = list_to_string(extracted_skills)

    # Convert the list to a single string
    converted_data = list_to_string(extracted_skills)
    # # print("\nfinal skills extracted in list are:",converted_data)
    # Create a new Excel sheet with the converted data
    create_excel_sheet([converted_data])
       
    
    # tokens = [token.text for token in nlp_text if not token.is_stop]
    
    required = job.Required_Skills
    match_percent = match_skills(converted_data,required)
    # print("Match Percentage:", match_percent)

    
    return match_percent, resume_data[titles[1]], resume_data[titles[2]], resume_data[titles[3]],  resume_data[titles[4]], resume_data[titles[5]]

#all the functions to be called
#pdf to text
def pdf_to_txt(pdf_path):
    resource_manager = PDFResourceManager(caching=True)
    
    # create a string object that will contain the final text the representation of the pdf. 
    out_text = StringIO()
    codec = 'utf-8'
    laParams = LAParams()
    
    # Create a TextConverter Object:
    text_converter = TextConverter(resource_manager, out_text, laparams=laParams)
    pdf_path = pdf_path.name 
    fp = open(pdf_path, 'rb')
    #Create a PDF interpreter object

    interpreter = PDFPageInterpreter(resource_manager, text_converter)
    
    # We are going to process the content of each page of the original PDF File    
    for page in PDFPage.get_pages(fp, pagenos=set(), maxpages=0, password="", caching=True, check_extractable=True):
        interpreter.process_page(page)

    
    # Retrieve the entire contents of the “file” at any time 
    text = out_text.getvalue()

    # Closing all the ressources we previously opened

    fp.close()
    text_converter.close()
    out_text.close()
    
    return text

# Extract txt from
def docx_to_txt(docx_path):
    temp = docx2txt.process(docx_path)
    text = [line.replace('\t', ' ') for line in temp.split('\n') if line]
    return ' '.join(text)

# detect file extension and call above functions accordingly
def extract_text(resume_file):
    # Access the resume file field from the model
    
    text = ''
    if resume_file.name.endswith('.pdf'):
        for page in pdf_to_txt(resume_file):
            text += '' + page
    elif resume_file.name.endswith('.docx') or resume_file.name.endswith('.doc'):
        text = docx_to_txt(resume_file)
    return text

# Filteration
def  filter_text(Text,lowerFlag=False,upperFlag=False,numberFlag=False,htmlFlag=False,urlFlag=False,punctFlag=False,spaceFlag=False,charFlag=False,bulletFlag=False,hashtagFlag=False,emojiFlag=False):
    if lowerFlag:
      Text = Text.lower()

    if upperFlag:
      Text = Text.upper()

    if numberFlag:
      import re
      Text = re.sub(r"\d+", '', Text)

    if htmlFlag:
      import re
      Text = re.sub(r'<[^>]*>', '', Text)

    if urlFlag:
      import re
      Text = re.sub(r'(https?|ftp|www)\S+', '', Text)

    if punctFlag:
      import re
      import string
      exclist = string.punctuation #removes [!”#$%&’()*+,-./:;<=>?@[\]^_`{|}~]
      # remove punctuations and digits from oldtext
      table_ = Text.maketrans('', '', exclist)
      Text = Text.translate(table_)

    if spaceFlag:
      import re
      Text = re.sub(' +'," ",Text).strip()

    if hashtagFlag:
      import re
      Text = re.sub(' +',"#",Text).strip()
      
    if charFlag:
      import re
      Text = re.sub('\n', ' ', Text)
      
    if bulletFlag:
      import re
      Text = re.sub('●','',Text)
      Text = re.sub('★','',Text)
      Text = re.sub('|','',Text)
      # Text = re.sub(([0-9|A-Z]+[\.|\)]|)\s+,'', Text)
    
    if emojiFlag:
      import emoji
      Text = emoji.sub(' +',"#",Text).strip()
      pass

    return Text

def extract_entity_sections(text):
    '''
    Helper function to extract all the raw text from sections of resume

    :param text: Raw text of resume
    :return: dictionary of entities
    '''
    RESUME_SECTIONS = [
                    'accomplishments',
                    'achievements',
                    'experience',
                    'education',
                    'interests',
                    'hobbies',
                    'projects',
                    'professional experience',
                    'experience',
                    'publications',
                    'skills',
                ]
    text_split = [i.strip() for i in text.split('\n')]
    # sections_in_resume = [i for i in text_split if i.lower() in sections]
    entities = {}
    key = False
    for phrase in text_split:
        if len(phrase) == 1:
            p_key = phrase
        else:
            p_key = set(phrase.lower().split()) & set(cs.RESUME_SECTIONS)
        try:
            p_key = list(p_key)[0]
        except IndexError:
            pass
        if p_key in RESUME_SECTIONS:
            entities[p_key] = []
            key = p_key
        elif key and phrase.strip():
            entities[key].append(phrase)
    
    # entity_key = False
    # for entity in entities.keys():
    #     sub_entities = {}
    #     for entry in entities[entity]:
    #         if u'\u2022' not in entry:
    #             sub_entities[entry] = []
    #             entity_key = entry
    #         elif entity_key:
    #             sub_entities[entity_key].append(entry)
    #     entities[entity] = sub_entities

    # # pprint.pprint(entities)

    # make entities that are not found None
    # for entity in cs.RESUME_SECTIONS:
    #     if entity not in entities.keys():
    #         entities[entity] = None 
    
    return entities

def extract_email(text):
    '''
    Helper function to extract email id from text

    :param text: plain text extracted from resume file
    '''
    email = re.findall("([^@|\s]+@[^@]+\.[^@|\s]+)", text)
    if email:
        try:
            return email[0].split()[0].strip(';')
        except IndexError:
            return None

def extract_mobile_number(text):
    '''
    Helper function to extract mobile number from text

    :param text: plain text extracted from resume file
    :return: string of extracted mobile numbers
    '''
    # Found this complicated regex on : https://zapier.com/blog/extract-links-email-phone-regex/
    phone = re.findall(re.compile(r'(?:(?:\+?([1-9]|[0-9][0-9]|[0-9][0-9][0-9])\s*(?:[.-]\s*)?)?(?:\(\s*([2-9]1[02-9]|[2-9][02-8]1|[2-9][02-8][02-9])\s*\)|([0-9][1-9]|[0-9]1[02-9]|[2-9][02-8]1|[2-9][02-8][02-9]))\s*(?:[.-]\s*)?)?([2-9]1[02-9]|[2-9][02-9]1|[2-9][02-9]{2})\s*(?:[.-]\s*)?([0-9]{4})(?:\s*(?:#|x\.?|ext\.?|extension)\s*(\d+))?'), text)
    if phone:
        number = ''.join(phone[0])
        if len(number) > 10:
            return '+' + number
        else:
            return number

def extract_name(text):
    # import spacy
    # from spacy.matcher import Matcher

    # nlp = spacy.load('en_core_web_sm')
    # # new_matcher = Matcher(nlp.vocab)
    # new_nlp = nlp(text)
    
    # pattern = [{'POS': 'PROPN'}, {'POS': 'PROPN'}, {'POS': 'PROPN'}]
    # matcher.add('NAME',[pattern], on_match=None)
    # matches = matcher(new_nlp)
    
    # for i, start, end in matches:
    #     span = new_nlp[start:end]
    #     # print("Name: ",span)
    tokenizer2 = AutoTokenizer.from_pretrained("Davlan/distilbert-base-multilingual-cased-ner-hrl")
    model2 = AutoModelForTokenClassification.from_pretrained("Davlan/distilbert-base-multilingual-cased-ner-hrl")
    nlp2 = pipeline("ner", model=model2, tokenizer=tokenizer2, aggregation_strategy="max")

    #example = """ Hello How are you doing I am fine Mrunal Kulkarni"""
    ner_results2 = nlp2(text)
    print(ner_results2)
    person_names = [entity['word'] for entity in ner_results2 if entity['entity_group'] == 'PER']
    print("YOUR NAME IS : ", person_names[0] if person_names else "No person name found")
    return person_names[0]

def list_to_string(lst):
    # Convert the list to a string representation
    return ', '.join(map(str, lst))

def extract_skills(nlp_text, noun_chunks):
    import os
    import pandas as pd

    file = 'skills.csv'
    # print("Current Working Directory:", os.getcwd())
    absolute_path = os.path.join(os.path.dirname(__file__), file)
    tokens = [token.text for token in nlp_text if not token.is_stop]

    data = pd.read_csv(absolute_path)
    skills = list(data.columns.values)
    skillset = []

    # check for one-grams
    for token in tokens:
        if token.lower() in skills:
            skillset.append(token)
    
    # check for bi-grams and tri-grams
    for token in noun_chunks:
        token = token.text.lower().strip()
        if token in skills:
            skillset.append(token)
            
    extracted_skills=[i.capitalize() for i in set([i.lower() for i in skillset])]
    # print(extracted_skills)
    return [i.capitalize() for i in set([i.lower() for i in skillset])]

def create_excel_sheet(data):
    # Specify the Excel file path
    excel_file = "webapp\Extracted Data\Extracted.xlsx"
    
    # Ensure the directory exists, create it if necessary
    os.makedirs(os.path.dirname(excel_file), exist_ok=True)

    # Load the existing Excel file
    try:
        wb = px.load_workbook(excel_file)
    except FileNotFoundError:
        # If the file doesn't exist, create a new workbook
        wb = px.Workbook()

    # # Select the worksheet where you want to append data (or create a new one)
    # ws = wb.active  # You can also select a specific sheet by name: workbook["SheetName"]

    # Get the current timestamp
    timestamp = datetime.now().strftime("%Y-%m-%d")
    # setect the current active sheet
    ws = wb.active
    # check if the current sheet has same timestamp
    if ws.title == timestamp:
        df = pd.DataFrame.from_dict([resume_data])
    
    else:
        # Create a new sheet with the timestamp as the name
        ws = wb.create_sheet(title=timestamp, index=0)
        ws["A1"] = 'Name'
        ws["B1"] = 'Email'
        ws["C1"] = 'Mobile No.'
        ws["D1"] = 'Eduaction'
        ws["E1"] = 'Skills'
        df = pd.DataFrame.from_dict([resume_data])
    
      
    for index, row in df.iterrows():
        ws.append(row.tolist())

    wb.save(excel_file)


# Matching the extracted skills with our requirements
def match_skills(skills,required):

    # # print("\nSkills in Resume: ",skills)
    # # print("\nRequired skills are: ",required)
    
    required_skills_list = [skill.strip() for skill in required.split(',')]
    # # print("\ntest rqd",required_skills_list)
    # print(required_skills_list)
    
    candi_skills_list = [skill.strip() for skill in skills.split(',')]
    # # print("\ntest rqd",candi_skills_list)
    # print(candi_skills_list)
    
    def dot_product(vec1, vec2):
        return sum(x * y for x, y in zip(vec1, vec2))

    def vector_length(vec):
        return sqrt(sum(x * x for x in vec))

    def cosine_similarity(vec1, vec2):
        if vector_length(vec1) == 0 or vector_length(vec2) == 0:
            return 0.0
        return dot_product(vec1, vec2) / (vector_length(vec1) * vector_length(vec2))
      
    lower_skill=[] 
    for skill in candi_skills_list:
        lower_skill.append(skill.lower())
        
    lower_req=[]
    for skill in required_skills_list:
        lower_req.append(skill.lower())
# Example usage:
    required_skills = lower_req
    candidate_skills = lower_skill
    # # print("\nrequired:",required_skills)
    # # print("\nskills:",candidate_skills)

# Convert skills to binary vectors
    required_skills_vector = [1 if skill in required_skills else 0 for skill in required_skills]
    candidate_skills_vector = [1 if skill in candidate_skills else 0 for skill in required_skills]
    matching=[]
    for skill in required_skills:
        if skill in candidate_skills:
            matching.append(skill)
    matched_skills= list_to_string(matching)
    # # print("\nmatched skills are:",matching)        
    similarity_percentage = cosine_similarity(required_skills_vector, candidate_skills_vector) * 100
    # # print("\nCandidate's skill fit percentage (using cosine similarity):", similarity_percentage)
    return similarity_percentage, matched_skills


def extract_education(nlp_text):
    patterns = [
        r'\b(?:BE|B\.E\.?|B\.S\.?|Bachelor of Engineering|Bachelor of Science)\b',
        r'\b(?:ME|M\.E\.?|M\.S\.?|Master of Engineering|Master of Science)\b',
        r'\b(?:BTECH|B\.TECH\.?|Bachelor of Technology)\b',
        r'\b(?:MTECH|M\.TECH\.?|Master of Technology)\b',
        r'\b(?:X|XII|HSC|SSC|H\.S\.C\.?|S\.C\.C\.?)\b',  # Include HSC, SSC, H.S.C, S.C.C
        r'\b(?:Diploma)\b',
    ]

    qualifications = []

    # Iterate through patterns and search for complete matches in the resume text
    for pattern in patterns:
        matches = re.findall(pattern, nlp_text)
        qualifications.extend(matches)

    # Remove duplicates and return the list of qualifications
    return list(set(qualifications))



# if __name__ == "__main__":
   
# import openpyxl
# from openpyxl import Workbook
# from datetime import datetime
# import pandas as pd

# def create_excel_sheet(dataframe):
#     # Create a new Excel workbook
#     wb = Workbook()
    
#     # Get the current timestamp for the sheet name
#     timestamp = datetime.now().strftime("%Y-%m-%d %H-%M-%S")
    
#     # Create a new sheet with the timestamp as the name
#     sheet = wb.create_sheet(title=timestamp)
    
#     # Convert the DataFrame to a list of lists for easy insertion
#     data = dataframe.values.tolist()
    
#     # Write the data to the Excel sheet
#     for row in data:
#         sheet.append(row)
    
#     # Save the Excel workbook
#     wb.save(f"output_{timestamp}.xlsx")

# if __name__ == "__main__":
#     # Sample DataFrame
#     data = {'Column1': [1, 2, 3, 4],
#             'Column2': ['A', 'B', 'C', 'D']}
#     df = pd.DataFrame(data)
    
#     # Create a new Excel sheet with the DataFrame
#     create_excel_sheet(df)








# from nltk.stem import WordNetLemmatizer 
# from nltk.corpus import stopwords

# def extract_experience(resume_text):
#     '''
#     Helper function to extract experience from resume text

#     :param resume_text: Plain resume text
#     :return: list of experience
#     '''
#     wordnet_lemmatizer = WordNetLemmatizer()
#     stop_words = set(stopwords.words('english'))

#     # word tokenization 
#     word_tokens = nltk.word_tokenize(resume_text)

#     # remove stop words and lemmatize  
#     filtered_sentence = [w for w in word_tokens if not w in stop_words and wordnet_lemmatizer.lemmatize(w) not in stop_words] 
#     sent = nltk.pos_tag(filtered_sentence)

#     # parse regex
#     cp = nltk.RegexpParser('P: {<NNP>+}')
#     cs = cp.parse(sent)
    
#     # for i in cs.subtrees(filter=lambda x: x.label() == 'P'):
#     #     # print(i)
    
#     test = []
    
#     for vp in list(cs.subtrees(filter=lambda x: x.label()=='P')):
#         test.append(" ".join([i[0] for i in vp.leaves() if len(vp.leaves()) >= 2]))

#     # Search the word 'experience' in the chunk and then # print out the text after it
#     x = [x[x.lower().index('experience') + 10:] for i, x in enumerate(test) if x and 'experience' in x.lower()]
#     return x
# # print(extract_experience(Text))




# exp_list=[]

# def extract_competencies(text, experience_list):
#     '''
#     Helper function to extract competencies from resume text
#     :param resume_text: Plain resume text
#     :return: dictionary of competencies
#     '''
#     experience_text = ' '.join(experience_list)
#     competency_dict = {}
#     COMPETENCIES = {
#     'teamwork': [
#         'supervised',
#         'facilitated',
#         'planned',
#         'plan',
#         'served',
#         'serve',
#         'project lead',
#         'managing',
#         'managed',
#         'lead ',
#         'project team',
#         'team',
#         'conducted',
#         'worked',
#         'gathered',
#         'organized',
#         'mentored',
#         'assist',
#         'review',
#         'help',
#         'involve',
#         'share',
#         'support',
#         'coordinate',
#         'cooperate',
#         'contributed'
#     ],
#     'communication': [
#         'addressed',
#         'collaborated',
#         'conveyed',
#         'enlivened',
#         'instructed',
#         'performed',
#         'presented',
#         'spoke',
#         'trained',
#         'author',
#         'communicate',
#         'define',
#         'influence',
#         'negotiated',
#         'outline',
#         'proposed',
#         'persuaded',
#         'edit',
#         'interviewed',
#         'summarize',
#         'translate',
#         'write',
#         'wrote',
#         'project plan',
#         'business case',
#         'proposal',
#         'writeup'
#     ],
#     'analytical': [
#         'process improvement',
#         'competitive analysis',
#         'aligned',
#         'strategive planning',
#         'cost savings',
#         'researched ',
#         'identified',
#         'created',
#         'led',
#         'measure',
#         'program',
#         'quantify',
#         'forecasr',
#         'estimate',
#         'analyzed',
#         'survey',
#         'reduced',
#         'cut cost',
#         'conserved',
#         'budget',
#         'balanced',
#         'allocate',
#         'adjust',
#         'lauched',
#         'hired',
#         'spedup',
#         'speedup',
#         'ran',
#         'run',
#         'enchanced',
#         'developed'
#     ],
#     'result_driven': [
#         'cut',
#         'decrease',
#         'eliminate',
#         'increase',
#         'lower',
#         'maximize',
#         'rasie',
#         'reduce',
#         'accelerate',
#         'accomplish',
#         'advance',
#         'boost',
#         'change',
#         'improve',
#         'saved',
#         'save',
#         'solve',
#         'solved',
#         'upgrade',
#         'fix',
#         'fixed',
#         'correct',
#         'achieve'           
#     ],
#     'leadership': [
#         'advise',
#         'coach',
#         'guide',
#         'influence',
#         'inspire',
#         'instruct',
#         'teach',
#         'authorized',
#         'chair',
#         'control',
#         'establish',
#         'execute',
#         'hire',
#         'multi-task',
#         'oversee',
#         'navigate',
#         'prioritize',
#         'approve',
#         'administer',
#         'preside',
#         'enforce',
#         'delegate',
#         'coordinate',
#         'streamlined',
#         'produce',
#         'review',
#         'supervise',
#         'terminate',
#         'found',
#         'set up',
#         'spearhead',
#         'originate',
#         'innovate',
#         'implement',
#         'design',
#         'launch',
#         'pioneer',
#         'institute'
#     ]
# }
#     for competency in COMPETENCIES.keys():
#         for item in COMPETENCIES[competency]:
#             if string_found(item, experience_text):
#                 if competency not in competency_dict.keys():
#                     competency_dict[competency] = [item]
#                 else:
#                     competency_dict[competency].append(item)
#     return competency_dict
# # print(extract_competencies(Text, exp_list))    



# def extract_measurable_results(text, experience_list):
#     '''
#     Helper function to extract measurable results from resume text

#     :param resume_text: Plain resume text
#     :return: dictionary of measurable results
#     '''

#     # we scan for measurable results only in first half of each sentence
#     experience_text = ' '.join([text[:len(text) // 2 - 1] for text in experience_list])
#     mr_dict = {}

#     for mr in cs.MEASURABLE_RESULTS.keys():
#         for item in cs.MEASURABLE_RESULTS[mr]:
#             if string_found(item, experience_text):
#                 if mr not in mr_dict.keys():
#                     mr_dict[mr] = [item]
#                 else:
#                     mr_dict[mr].append(item)
    
#     return mr_dict



# # ner using xlnet testing 



# import torch
# from transformers import XLNetForSequenceClassification, XLNetTokenizer, AdamW
# from torch.utils.data import DataLoader, Dataset
# import json

# # Load your JSON dataset
# with open("traindata3.json", "r", encoding="utf-8") as json_file:
#     data = json.load(json_file)




# data[0]





# def process_label_and_points(label, points):
#     # Initialize an empty list to store the extracted entity texts
#     entity_texts = []
    
#     # Loop through each point dictionary in the list
#     for point in points:
#         start = point['start']
#         end = point['end']
        
#         # Extract the labeled entity text from the content
#         entity_text = content[start:end]
        
#         # Append the entity text to the list
#         entity_texts.append(entity_text)
    
#     # Combine the extracted entity texts if needed (e.g., join with a space)
#     combined_entity_text = ' '.join(entity_texts)
    
#     return combined_entity_text





# input_data




# from transformers import XLNetTokenizer
# import torch

# # Initialize the XLNet tokenizer
# tokenizer = XLNetTokenizer.from_pretrained('xlnet-base-cased')

# # Tokenize and pad/truncate input sequences
# input_ids = []
# attention_masks = []

# for text in Text:
#     encoded_text = tokenizer.encode_plus(
#         text,
#         add_special_tokens=True,
#         max_length=128,  # Adjust the max length as needed
#         padding='max_length',
#         return_attention_mask=True,
#         return_tensors='pt'
#     )

#     input_ids.append(encoded_text['input_ids'])
#     attention_masks.append(encoded_text['attention_mask'])

# # Convert input_ids and attention_masks to tensors
# input_ids = torch.cat(input_ids, dim=0)
# attention_masks = torch.cat(attention_masks, dim=0)

# # Encode the target labels (NER tags) into tensors or other suitable format
# # You may need to use label encoders or other methods based on your model's requirements



# input_ids




# import torch
# from torch.utils.data import Dataset

# class NERDataset(Dataset):
#     def __init__(self, input_ids, attention_masks, target_labels):
#         """
#         Args:
#             input_ids (list of torch.Tensor): List of input sequences as tensors.
#             attention_masks (list of torch.Tensor): List of attention masks as tensors.
#             target_labels (list of list of str): List of target labels for NER, where each list contains labels for one example.
#         """
#         self.input_ids = input_ids
#         self.attention_masks = attention_masks
#         self.target_labels = target_labels

#     def __len__(self):
#         return len(self.input_ids)

#     def __getitem__(self, idx):
#         sample = {
#             'input_ids': self.input_ids[idx],
#             'attention_mask': self.attention_masks[idx],
#             'target_labels': self.target_labels[idx],
#         }
#         return sample




# from transformers import XLNetForTokenClassification, Trainer, TrainingArguments
# label_list = ['College Name', 'Designation', 'Name', 'Companies worked at', 'Email Address', 'Location', 'Skills']
# target_labels = [
#     ["O", "O", "B-ORG", "I-ORG", "O", ],  # Labels for the first example
#     ["B-PER", "I-PER", "O", "B-LOC", "O", ]  # Labels for the second example
#     # Add labels for more examples...
# ]

# # Calculate the number of unique labels
# num_labels = len(label_list)# Load the pre-trained XLNet model for token classification
# model = XLNetForTokenClassification.from_pretrained('xlnet-base-cased', num_labels=num_labels)  # Define num_labels accordingly

# # Define training arguments (batch size, learning rate, etc.)
# training_args = TrainingArguments(
#     output_dir='./ner_model',
#     per_device_train_batch_size=16,
#     num_train_epochs=3,
#     learning_rate=2e-5,
# )

# # Initialize the Trainer
# trainer = Trainer(
#     model=model,
#     args=training_args,
#     train_dataset=NERDataset(input_ids, attention_masks, target_labels),  # Define your custom dataset class
# )

# # Start training
# trainer.train()




# pip install transformers[torch]




# pip install accelerate -U







